from flask import Flask, render_template, request
from datetime import datetime
from fileinput import filename
from pymongo.mongo_client import MongoClient
from pymongo.server_api import ServerApi
from flask_cors import CORS
import openpyxl

app = Flask(__name__, template_folder='pages')
CORS(app)
uri = "mongodb+srv://unurag:89800932@cluster0.tjnuduv.mongodb.net/?retryWrites=true&w=majority&appName=Cluster0"
client = MongoClient(uri, server_api=ServerApi('1'))
db = client["eventname"]



@app.route('/')
def index():
    return "Not allowed", 401

# open a page where main xlsx file can be uploaded
@app.route('/secret/uploadpdf')
def upload():
    return render_template('upload.html')

@app.route('/secret/filesuc', methods = ['POST'])
def filesuc():
    if request.method == 'POST':
        f = request.files['file']
        f.save(f.filename)
        return f"File {f.filename} uploaded successfully"

    return "File uploaded succesfully"

# creating an event
@app.route('/secret/createevent')
def createEvent():
    return render_template('cevent.html')

@app.route('/secret/eventsuc', methods = ['POST'])
def createEventSuc():
    if request.method == 'POST':
        eventName = request.form.get('eventname')
        startDate = request.form.get('startdate')
        endDate = request.form.get('enddate')
        objStartDate = datetime.strptime(startDate,"%Y-%m-%d")
        objEndDate = datetime.strptime(endDate,"%Y-%m-%d")

        collection = db["event"]
        post = {"_id": 1, "eventname": eventName, "startdate": f"{objStartDate.day}/{objStartDate.month}/{objStartDate.year}", "enddate": f"{objEndDate.day}/{objEndDate.month}/{objEndDate.year}"}
        collection.delete_many({})
        collection.insert_one(post)

        return 'Event created successfully'
    
@app.route('/secret/resultsuc', methods = ['POST'])
def createResult():
    if request.method == 'POST':
        weight = request.form.get('weight')
        gender = request.form.get('gender')
        athleteOne = request.form.get('athleteone')
        stateOne = request.form.get('statecodeone')
        sideOne = request.form.get('sideone')
        scoreOne = request.form.get('athleteonescore')
        athleteTwo = request.form.get('athletetwo')
        stateTwo = request.form.get('statecodetwo')
        sideTwo = request.form.get('sidetwo')
        scoreTwo = request.form.get('athletetwoscore')
        winner = request.form.get('winner')

        collection = db["results"]
        query = {"_id": weight}
        document = collection.find_one(query)
        
        num = 0 if document is None else len(document)

        post = {"_id": weight, 
                    "1": {
                    "nameone": athleteOne,
                    "gender": gender,
                    "stateone": stateOne,
                    "sideone": sideOne,
                    "scoreone": scoreOne,
                    "statetwo": stateTwo,
                    "nametwo": athleteTwo,
                    "sidetwo": sideTwo,
                    "scoretwo": scoreTwo,
                    "winner": winner
                    }
                }

        if num == 0 or collection.count_documents({}) == 0:
            collection.insert_one(post)
        else:
            postupdate = {str(num): {
                        "nameone": athleteOne,
                        "gender": gender,
                        "stateone": stateOne,
                        "sideone": sideOne,
                        "scoreone": scoreOne,
                        "statetwo": stateTwo,
                        "nametwo": athleteTwo,
                        "sidetwo": sideTwo,
                        "scoretwo": scoreTwo,
                        "winner": winner
                    }}
            collection.update_one(query, {"$set": postupdate})
        
    return "Result has been declared.."
@app.route('/secret/resultupload')
def result():
    return render_template('resultform.html')

@app.route('/results')
def res():
    collection = db["results"]
    init = collection.find()
    
    for docs in init:
        print(docs)

    return 'data'

@app.route('/event')
def event():
    collection = db['event']
    eventDet = collection.find_one({"_id":1})
    return eventDet

@app.route('/formdatafetch')
def formfetch():
    try:
        xl_file = openpyxl.load_workbook('./new.xlsx')
        sheet = xl_file.active

        dictn = {}

        total_rows = sheet.max_row

        for row in sheet.iter_rows(min_row=2, values_only=True):
            state = row[0]
            name = row[1]
            weight = row[2]

            if weight in dictn:
                dictn[weight][0].append(state)
                dictn[weight][1].append(name)
            else:
                dictn[weight] = ([state], [name])
    except FileNotFoundError:
        return "Error: File not found."
    except Exception as e:
        return "An error occurred: " + str(e)
    
    return render_template('resultform.html', data=dictn)

if __name__ == "__main__":
    app.run(debug=True)
    CORS(app)